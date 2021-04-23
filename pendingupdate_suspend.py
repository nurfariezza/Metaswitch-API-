import traceback, smtplib, utils, os, shutil,time, datetime
from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
from datetime import datetime
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
from models.clientlist import Clientlist
from utils import getConnection
from os import listdir
from os.path import isfile, join
from datetime import date, datetime
from dateutil.relativedelta import relativedelta



def sendMail(output):
    strfrom = 'redtonernd@biz.redtone.com'
    tolist =  ['ainur.fadzil@redtone.com']
    cclist = ['ainur.fadzil@redtone.com']
    
    strto = ', '.join(tolist)
    strcc = ', '.join(cclist)
    
    msg = MIMEMultipart()
    msg['Subject'] = 'Failed Sync to Metaswitch ' 
    msg['From'] = 'Metaswitch Service <{0}>'.format(strfrom)
    msg['To'] = strto
    msg['Cc'] = strcc

    part = MIMEBase('application', "vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    part.set_payload(output)
    encoders.encode_base64(part)

    part.add_header('Content-Disposition', 'attachment; filename="Metaswitch Service-{0}.xlsx"'.format(datetime.now().strftime('%Y-%m-%d')))

    msg.attach(part)

    server = smtplib.SMTP(host='mail.redtone.com', port=587)
    server.sendmail(strfrom, tolist + cclist, msg.as_string())
    print('Sent message successfully....')

def sendfile(output,IP):
    dt = datetime.now()

    path = 'output_list'
    filenameloc = path
    filen = IP
    if not os.path.exists(filenameloc):
        os.makedirs(filenameloc)
   
    filename = str(filen)+'_{0}_Metaswitch Service List.xlsx'.format(dt.strftime('%m%d%Y'))
    with open(os.path.join(path, filename), 'wb') as temp_file:
        temp_file.write(output)

def getData(l):
    wb = Workbook()
    ws = wb.active
    
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15

    
    headers = ["", "SnapPicture", "Name", "Number", "ListMode", "Temperature", "Time", "Information"]
    for i in range(0, len(headers)):
        j = i + 1
        ws.cell(row=1, column=j).value = headers[i]
        
    i = 2
    for o in l:
        j = 1
        cell = ws.cell(row=i, column=j)
        cell.value = ''
        ws.row_dimensions[i].height = 50

        
        j = j + 1

        path = o.Path
        path = open(os.path.join(path, o.Snapname), 'rb')
        
        img = Image(path) 

        im = PILImage.open(path)
        pil_img = im.resize((80, 60))
        pil_img.save(o.Snapname)
        img = Image(o.Snapname)

        imgcell= 'B{0}'.format(i)
        ws.add_image(img, imgcell)
        
        j = j + 1
        ws.cell(row=i, column=j).value = o.name
        
        j = j + 1
        ws.cell(row=i, column=j).value = o.userid
        
       
        j = j + 1
        ws.cell(row=i, column=j).value = o.face_mode

        j = j + 1
        o.Temperature = "{:.2f}".format(o.Temperature)
        ws.cell(row=i, column=j).value = o.Temperature
        
        j = j + 1
        ws.cell(row=i, column=j).value = o.time
        cell.number_format = 'yyyy-mm-dd'

        j = j + 1
        ws.cell(row=i, column=j).value ='{} {} {} {}'.format(o.Similarity, o.guest, o.mask, o.Temp_read)
        #ws.cell(row=i, column=j).value = ' '.join(o.glasses)

       
        i = i + 1
        
    b = save_virtual_workbook(wb)
    return b

def getIP(cur):
    l = []
  
    try:
        
        q = """
            select deviceIP from CmpInfoList group by deviceIP
            """
        cur.execute(q)
        rows = cur.fetchall()
        for x in rows:
            #m = Clientlist()
            deviceIP = x[0]
            l.append(deviceIP)
            #print(deviceIP)

    except Exception:
        print(traceback.format_exc())
        
    return l

def getCustomer(cur, dxn, dtn, IP):
#def getCustomer(cur, IP):
    
    l = []
  
    try:
        print(IP)
        #IP=str(IP).strip("[]")
        IP= str(IP).replace('[','').replace(']','')
        q = """
            select Path, libname, Snapname, name, userid, face_mode, Temperature, time, Similarity, deviceIP, guest  ,glasses, mask  from CmpInfoList
            where time between ? and ? and deviceIP = ?
            """

        IP= '{0}'.format(str(IP))         
        

        params = (dxn, dtn, '192.168.1.88')
        #params= str(params).replace('"','').replace('"','')

        print(params) 
        rows = cur.execute(q, params).fetchall()
        
        #rows = cur.execute("select Path, libname, Snapname, name, userid, face_mode, Temperature, time, Similarity, deviceIP, guest  ,glasses, mask  from CmpInfoList where time between ? and ? and deviceIP = ?;", str(str(dxn), str(dtn), str(IP)))
        
        print(IP)

        #rows = cur
        #print(rows)
        for r in rows:
            o = Clientlist()
            #print(o)
            o.Path = r[0] 
            if o.Path == '': 
                o.Path == 'N/A'
            o.libname = r[1]
            if o.libname == '': 
                o.libname == 'N/A'

            o.Snapname = r[2]
            if o.Snapname == '': 
                o.Snapname == 'N/A'

            o.name = r[3]
            if o.name == '': 
                o.name == 'N/A'

            o.userid = r[4]
            if o.userid == '': 
                o.userid == 'N/A'

            o.face_mode = r[5]
            if o.face_mode == 0:
                o.face_mode='Stranger'
            elif o.face_mode == 1:
                o.face_mode='Blacklist'
            elif o.face_mode == 2:
                o.face_mode='Whitelist'
            elif o.face_mode == 3:
                o.face_mode='VIP List'


            o.Temperature = r[6]
           
            if o.Temperature <= float(o.Temperature):
                o.Temp_read='Temperature is all right'
            else:
                o.Temp_read='Temperature is high'

            o.time = r[7]
            o.Similarity = r[8] #"{:.2f}".format(a_float)
            if o.Similarity is not None:
                o.Similarity = "{:.2f}".format(o.Similarity*100)#float(o.Similarity*100)
                o.Similarity='Similarity: {}'.format(o.Similarity)
            else:
                o.Similarity='Similarity:N/A'
            o.deviceIP = r[9]

            o.guest = r[10]
            if o.guest is not None:
                o.guest='Visit Times: {}'.format(o.guest)
            else:
                o.guest='Visit Times:N/A'
            o.glasses = r[11]
            
                
            o.mask = r[12]
            if o.mask == 1:
                o.mask='Mask:Yes,'
            else:
                o.mask='Mask:No,'
            
            l.append(o)

         

            
    except Exception:
        print(traceback.format_exc())
        
    return l

 

if __name__ == '__main__':
    
    con = None
    cur = None
    #destpath = 'output'
    
    
    try:
        dt = datetime.now()
        dtn = dt + relativedelta(days=+1)
        dtn = dtn.strftime('%Y-%m-%d')
        #print(dtn)
        dx = dt + relativedelta(days=-15)
        dxn = dx.strftime('%Y-%m-%d')

        #print(dxn)
        con, cur = getConnection() 
        IP = getIP(cur)

        l = getCustomer(cur, dxn, dtn, IP)
        #l = getCustomer(cur,  IP)

        #getData(l)
        output = getData(l)
        #sendfile(output,IP)


       
    
    except Exception:
        print(traceback.format_exc())
        
    finally:
        if cur is not None:
            cur.close()
            
        if con is not None:
            con.close()

    
# ######
   
